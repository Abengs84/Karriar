from dataclasses import dataclass
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Student


def _cell(row, idx: int) -> str | None:
    val = row[idx] if idx < len(row) else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _capitalize_name_part(part: str) -> str:
    """Ett ord eller bindestrecksdelat namn: 'cajsa' / 'anna-maria' -> 'Cajsa' / 'Anna-Maria'."""
    return "-".join(p.capitalize() for p in part.split("-"))


def capitalize_person_name(name: str) -> str:
    """För- eller efternamn: 'cajsa maris' -> 'Cajsa Maris'."""
    return " ".join(_capitalize_name_part(part) for part in name.split())


_TIMESTAMP_FORMATS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
)


def _parse_timestamp(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in _TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


@dataclass
class _ParsedRow:
    row: tuple
    first_name: str
    last_name: str
    school: str
    key: tuple[str, str, str]
    timestamp: datetime | None
    file_order: int


def _prefer_newer_row(candidate: _ParsedRow, current: _ParsedRow) -> bool:
    """True om candidate ska ersätta current (senare timestamp eller senare rad i filen)."""
    if candidate.timestamp and current.timestamp:
        if candidate.timestamp != current.timestamp:
            return candidate.timestamp > current.timestamp
    elif candidate.timestamp:
        return True
    elif current.timestamp:
        return False
    return candidate.file_order > current.file_order


def _dedupe_rows_by_student(rows: list[_ParsedRow]) -> tuple[list[_ParsedRow], int]:
    """Behåll senaste anmälan per elev (namn + skola) utifrån timestamp."""
    best: dict[tuple[str, str, str], _ParsedRow] = {}
    for parsed in rows:
        prev = best.get(parsed.key)
        if prev is None or _prefer_newer_row(parsed, prev):
            best[parsed.key] = parsed
    deduped = list(best.values())
    skipped = len(rows) - len(deduped)
    return deduped, skipped


def import_students_from_excel(db: Session, file_bytes: bytes) -> tuple[int, int]:
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    imported = 0
    skipped = 0
    rows = ws.iter_rows(values_only=True)

    first = next(rows, None)
    if first is None:
        return 0, 0

    # Skip header if column B or C looks like a name label
    name_header_labels = (
        "förnamn", "fornamn", "firstname", "etunimi",
        "efternamn", "efternamn", "lastname", "sukunimi",
    )
    col_b = str(first[1]).lower().strip() if first[1] else ""
    col_c = str(first[2]).lower().strip() if len(first) > 2 and first[2] else ""
    if col_b in name_header_labels or col_c in name_header_labels:
        pass
    else:
        rows = iter([first, *rows])

    existing_keys = {
        (s.first_name.lower(), s.last_name.lower(), s.school.lower())
        for s in db.query(Student).all()
    }

    parsed_rows: list[_ParsedRow] = []
    for file_order, row in enumerate(rows):
        if not row or len(row) < 4:
            continue
        last_name = _cell(row, 1)
        first_name = _cell(row, 2)
        school = _cell(row, 3)
        if not first_name or not last_name or not school:
            continue

        first_name = capitalize_person_name(first_name)
        last_name = capitalize_person_name(last_name)
        key = (first_name.lower(), last_name.lower(), school.lower())
        parsed_rows.append(
            _ParsedRow(
                row=row,
                first_name=first_name,
                last_name=last_name,
                school=school,
                key=key,
                timestamp=_parse_timestamp(row[0] if row else None),
                file_order=file_order,
            )
        )

    deduped_rows, skipped_in_file = _dedupe_rows_by_student(parsed_rows)
    skipped += skipped_in_file

    for parsed in deduped_rows:
        if parsed.key in existing_keys:
            skipped += 1
            continue

        student = Student(
            first_name=parsed.first_name,
            last_name=parsed.last_name,
            school=parsed.school,
            choice1=_cell(parsed.row, 4),
            choice2=_cell(parsed.row, 5),
            choice3=_cell(parsed.row, 6),
            reserve=_cell(parsed.row, 7),
        )
        db.add(student)
        existing_keys.add(parsed.key)
        imported += 1

    db.commit()
    wb.close()
    return imported, skipped
