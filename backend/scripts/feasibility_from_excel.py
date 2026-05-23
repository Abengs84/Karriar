"""
Genomförbarhetsanalys utifrån Responses.xlsx och Karriär-rummen.

  python backend/scripts/feasibility_from_excel.py [sökväg-till-xlsx]

Kör CP-SAT med samma regler som global placering (min 5 elever/session,
ett rum per inspiratör / hybrid, lunchbalans).
"""

from __future__ import annotations

import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from openpyxl import load_workbook  # noqa: E402

from app.auto_placer import RoomRef, StudentRef  # noqa: E402
from app.import_excel import capitalize_person_name  # noqa: E402
from app.placement_cp_sat import CpSatConfig, solve_cp_sat  # noqa: E402
from app.seed_rooms import KARRIAR_CONFIRMED_ROOMS  # noqa: E402


def _cell(row, idx: int) -> str | None:
    val = row[idx] if idx < len(row) else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def load_students_from_excel(path: Path) -> list[StudentRef]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    first = rows[0]
    col_b = str(first[1]).lower().strip() if first[1] else ""
    data_rows = rows[1:] if col_b in ("efternamn", "fornamn", "förnamn") else rows

    best: dict[tuple[str, str, str], tuple[datetime | None, tuple]] = {}
    for file_order, row in enumerate(data_rows):
        if not row or len(row) < 4:
            continue
        last_name = _cell(row, 1)
        first_name = _cell(row, 2)
        school = _cell(row, 3)
        if not first_name or not last_name or not school:
            continue
        first_name = capitalize_person_name(first_name)
        last_name = capitalize_person_name(last_name)
        key = (first_name.lower(), last_name.lower(), school.lower())
        ts = row[0] if isinstance(row[0], datetime) else None
        prev = best.get(key)
        if prev is None or (ts and prev[0] and ts > prev[0]) or (ts and not prev[0]):
            best[key] = (ts, row)

    students: list[StudentRef] = []
    for i, (_ts, row) in enumerate(best.values(), start=1):
        students.append(
            StudentRef(
                id=i,
                choice1=_cell(row, 4),
                choice2=_cell(row, 5),
                choice3=_cell(row, 6),
                reserve=_cell(row, 7),
                lunch_track=None,
                placements=[],
            )
        )
    return students


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "Responses.xlsx"
    if not xlsx.is_file():
        print(f"Hittar inte {xlsx}")
        sys.exit(1)

    students = load_students_from_excel(xlsx)
    rooms = [
        RoomRef(i + 1, name, cap)
        for i, (name, cap) in enumerate(KARRIAR_CONFIRMED_ROOMS)
    ]
    total_cap = sum(r.capacity for r in rooms)

    demand: Counter[str] = Counter()
    for s in students:
        seen: set[str] = set()
        for field in ("choice1", "choice2", "choice3"):
            val = getattr(s, field)
            if val and val not in seen:
                demand[val] += 1
                seen.add(val)

    print(f"Excel: {xlsx.name}")
    print(f"Elever (dedup): {len(students)}")
    print(f"Rum (seed): {len(rooms)}, summa platser per tid: {total_cap}")
    print(f"Inspiratörer med val 1–3: {len(demand)}")
    print(f"Behov placeringar (3/elev): {len(students) * 3}")
    print()
    print("Topp 8 efterfrågan:")
    for insp, n in demand.most_common(8):
        print(f"  {n:3d}  {insp[:60]}")
    print()

    overflow = len(demand) - len(rooms)
    if overflow > 0:
        print(
            f"OBS: {overflow} fler inspiratörer an rum – hybrid kravs for minst valda."
        )
    if total_cap < len(students):
        print(
            f"OBS: Platser per tid ({total_cap}) < elever ({len(students)})."
        )
    print()
    print("Kör CP-SAT (kan ta upp till 2 minuter)...")

    slots: list = []
    _, _, result, diag = solve_cp_sat(
        students,
        rooms,
        slots,
        config=CpSatConfig(
            min_session_size=5,
            time_limit_seconds=180.0,
            balance_lunch_tracks=True,
            same_room_per_inspirator=True,
            hybrid_room_when_short=True,
        ),
    )

    print()
    print(f"Status: {diag.status} ({diag.wall_time:.1f}s)")
    print(f"Placerade val 1–3: {diag.placed_required} / {diag.required_total}")
    print(f"Oplacerade val: {len(result.unplaced_needs)}")
    print(f"Poäng: {result.score}")
    print(f"Sammanfattning: {result.summary}")
    if diag.infeasible_hints:
        print("Tips:")
        for h in diag.infeasible_hints:
            print(f"  - {h}")
    if diag.small_sessions:
        print("Små sessioner:")
        for insp, pt, n in diag.small_sessions[:10]:
            print(f"  - {insp} {pt}: {n} elever")


if __name__ == "__main__":
    main()
